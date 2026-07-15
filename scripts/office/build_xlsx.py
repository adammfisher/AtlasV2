#!/usr/bin/env python3
"""Compile table-model JSON (skills/xlsx/schema.json) via openpyxl.

Doctrine: named styles only; explicit number format on every numeric column;
frozen header row; real Excel Table with the spec's named table style;
content-approximated column widths; print area set; formulas stay formulas
(the spec gate already rejected hardcoded derived rows); financial color code
blue=input / black=formula / green=link / red=external applied as named styles.
"""
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.styles import Font, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import validate_common as vc

BRAND_DARK = "371447"
TEXT_DARK = "1B1A18"
INPUT_BLUE = "0B5394"    # financial-model input convention
LINK_GREEN = "006100"
EXTERNAL_RED = "9C0006"
FONT = "Helvetica Neue"

NUMBER_FORMATS = {
    "text": "@",
    "integer": "#,##0",
    "decimal": "#,##0.00",
    "currency": '"$"#,##0',
    "percent": "0.0%",
    "date": "yyyy-mm-dd",
}


def _named_styles(wb):
    def add(name, **font_kw):
        style = NamedStyle(name=name)
        style.font = Font(name=FONT, **font_kw)
        if name == "atlas_header":
            style.fill = PatternFill("solid", fgColor=BRAND_DARK)
        wb.add_named_style(style)

    add("atlas_header", bold=True, color="FFFFFF")
    add("atlas_input", color=INPUT_BLUE)
    add("atlas_formula", color=TEXT_DARK)
    add("atlas_link", color=LINK_GREEN)
    add("atlas_external", color=EXTERNAL_RED)
    add("atlas_text", color=TEXT_DARK)


def _cell_style(cell_spec, column_role) -> str:
    """Financial color code: content decides, column role refines."""
    if column_role == "link":
        return "atlas_link"
    if column_role == "external":
        return "atlas_external"
    if isinstance(cell_spec, dict) and "f" in cell_spec:
        return "atlas_formula"
    if isinstance(cell_spec, dict) and "n" in cell_spec:
        return "atlas_input"
    return "atlas_text"


def build(payload: dict, template: str | None, out: Path) -> dict:
    wb = Workbook()
    wb.remove(wb.active)
    _named_styles(wb)
    cells = formulas = 0
    for sheet_index, sheet_spec in enumerate(payload["sheets"]):
        ws = wb.create_sheet(title=str(sheet_spec["name"])[:31])
        columns = sheet_spec["columns"]
        rows = sheet_spec["rows"]
        ncol = len(columns)

        # header row: named style + frozen
        for c, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=c, value=str(col["header"]))
            cell.style = "atlas_header"
        ws.freeze_panes = "A2"

        widths = [len(str(col["header"])) for col in columns]
        for r, row in enumerate(rows, start=2):
            for c in range(1, ncol + 1):
                spec = row[c - 1] if c - 1 < len(row) else None
                cell = ws.cell(row=r, column=c)
                if spec is None:
                    continue
                if "f" in spec:
                    cell.value = spec["f"]
                    formulas += 1
                elif "n" in spec:
                    cell.value = spec["n"]
                else:
                    cell.value = spec.get("t", "")
                col_spec = columns[c - 1]
                cell.style = _cell_style(spec, col_spec.get("role"))
                cell.number_format = NUMBER_FORMATS[col_spec["format"]]
                rendered = str(spec.get("t", spec.get("n", spec.get("f", ""))))
                widths[c - 1] = max(widths[c - 1], min(len(rendered), 38))
                cells += 1

        # column widths: max-content-length approximation
        for c, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = max(9, min(40, width * 1.15 + 2))

        # a real Excel table carries banding + the spec's named style
        ref = f"A1:{get_column_letter(ncol)}{1 + len(rows)}"
        table = Table(displayName=f"AtlasTable{sheet_index + 1}", ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name=sheet_spec["table_style"], showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)

        ws.print_area = ref
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        if ncol > 6:
            ws.page_setup.orientation = "landscape"

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return {"sheets": len(payload["sheets"]), "cells": cells, "formulas": formulas, "bytes": out.stat().st_size}


def main() -> None:
    args = vc.cli("xlsx builder")
    payload = vc.load_payload(args.payload)
    vc.spec_gate("xlsx", payload)
    out = Path(args.out)
    meta = build(payload, args.template, out)

    checks = [vc.openxml_audit(out), vc.zip_sanity(out)]

    reopened = load_workbook(str(out))
    checks.append(vc.check("Round-trip", len(reopened.sheetnames) == meta["sheets"]))

    # formula syntax via openpyxl tokenizer
    syntax_ok = True
    frozen_ok = True
    formats_ok = True
    texts = []
    for ws in reopened.worksheets:
        if ws.freeze_panes != "A2":
            frozen_ok = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, str):
                    texts.append(cell.value)
                    if cell.value.startswith("="):
                        try:
                            Tokenizer(cell.value)
                        except Exception:
                            syntax_ok = False
                if isinstance(cell.value, (int, float)) and cell.number_format == "General":
                    formats_ok = False
    checks.append(vc.check("Formula syntax", syntax_ok))
    checks.append(vc.check("Header row frozen", frozen_ok))
    checks.append(vc.check("Number formats on numeric cells", formats_ok))
    checks.append(vc.placeholder_grep(texts))
    checks.append(vc.soffice_recalc_scan(out))
    vc.emit(out, meta, checks)


if __name__ == "__main__":
    try:
        main()
    except SystemExit:
        raise
    except Exception as err:
        vc.fail(f"{type(err).__name__}: {err}")
